import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Configuration parameters
INPUT_XLSX = 'Resources.xlsx'
DOWNLOAD_DIR = 'Resources'
WEBSITE_URL = 'https://libgen.rs/'  # Replace with the given website URL

# Ensure download directory exists
if not os.path.exists(DOWNLOAD_DIR):
    os.makedirs(DOWNLOAD_DIR)

def construct_search_url(query):
    """
    Constructs a search URL for the target website.
    This function assumes the website supports queries via a GET parameter, e.g. ?q=...
    Adjust according to the actual website search format.
    """
    return f"{WEBSITE_URL}/search?q={requests.utils.quote(query)}"

def parse_search_results(html, book_title):
    """
    Parses the HTML search results and tries to locate a download link that matches the book title.
    Adjust the parsing logic based on the website's actual HTML structure.
    """
    soup = BeautifulSoup(html, 'html.parser')
    # Example: assume each result is in a div with class 'result'
    for result in soup.find_all('div', class_='result'):
        # Assume title is in an <h2> tag within the result
        title_tag = result.find('h2')
        if title_tag and book_title.lower() in title_tag.get_text(strip=True).lower():
            # Assume the download link is in an <a> tag with class 'download'
            download_link_tag = result.find('a', class_='download')
            if download_link_tag and download_link_tag.get('href'):
                # Build absolute URL if needed
                link = download_link_tag['href']
                if not link.startswith('http'):
                    link = WEBSITE_URL + link
                return link
    return None

def download_file(url, save_path):
    """
    Downloads a file from a URL and saves it to the specified path.
    Returns True if the download succeeds, False otherwise.
    """
    try:
        response = requests.get(url, stream=True, timeout=15)
        response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception as e:
        print(f"Download failed from {url}: {e}")
        return False

def search_and_download(book_title):
    """
    Searches for the book title on the website and attempts to download it.
    If the exact book is not found, it tries searching for an older edition.
    Returns True if a download is successful.
    """
    # First try: search for the exact title
    search_query = book_title
    search_url = construct_search_url(search_query)
    print(f"Searching for: {search_query}")
    try:
        response = requests.get(search_url, timeout=15)
        response.raise_for_status()
    except Exception as e:
        print(f"Error searching for {book_title}: {e}")
        return False

    download_link = parse_search_results(response.text, book_title)
    
    # Second try: if not found, search for older editions
    if not download_link:
        search_query = f"{book_title} older edition"
        search_url = construct_search_url(search_query)
        print(f"Exact title not found. Searching for older editions with query: {search_query}")
        try:
            response = requests.get(search_url, timeout=15)
            response.raise_for_status()
        except Exception as e:
            print(f"Error searching for older editions of {book_title}: {e}")
            return False
        download_link = parse_search_results(response.text, book_title)
    
    if not download_link:
        print(f"Download link not found for {book_title}.")
        return False
    
    # Determine a filename from the book title (sanitize it for filesystem)
    safe_title = "".join([c if c.isalnum() or c in " ._-" else "_" for c in book_title])
    file_path = os.path.join(DOWNLOAD_DIR, f"{safe_title}.pdf")  # Assuming PDF format; adjust if needed
    
    print(f"Downloading {book_title} from {download_link}")
    if download_file(download_link, file_path):
        print(f"Downloaded {book_title} successfully.")
        return True
    else:
        print(f"Failed to download {book_title}.")
        return False

def update_workbook(download_status):
    """
    Reads the XLSX file, updates the 'Downloaded' column for each book based on download_status,
    and saves the workbook.
    download_status is a dictionary mapping row numbers (or book titles) to a boolean.
    """
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    # Identify header indices based on the first row
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    if 'Title' not in headers or 'Downloaded' not in headers:
        print("XLSX file must contain 'Title' and 'Downloaded' columns.")
        return

    # Update each row: assume Title is unique and use it as key in download_status dict.
    for row in ws.iter_rows(min_row=2):
        title_cell = row[headers['Title'] - 1]
        downloaded_cell = row[headers['Downloaded'] - 1]
        if title_cell.value and download_status.get(title_cell.value, False):
            downloaded_cell.value = "Yes"

    wb.save(INPUT_XLSX)
    print("Workbook updated with download statuses.")

def main():
    # Load the workbook and build a list of titles to process
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    # Identify header indices based on the first row
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    if 'Title' not in headers or 'Downloaded' not in headers:
        print("XLSX file must contain 'Title' and 'Downloaded' columns.")
        return

    download_status = {}  # Dictionary to keep track of download status per book title

    # Iterate over the rows and process each book/article
    for row in ws.iter_rows(min_row=2):
        title = row[headers['Title'] - 1].value
        downloaded = row[headers['Downloaded'] - 1].value
        # Skip if already downloaded
        if downloaded and str(downloaded).strip().lower() == "yes":
            continue
        if not title:
            continue

        success = search_and_download(title)
        download_status[title] = success

    # Update the XLSX file with the results
    update_workbook(download_status)

if __name__ == '__main__':
    main()
